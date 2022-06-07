# print("Hesam")
# print("*" * 10)
# age = input("What is your age? ")
# print('You are ' + age)
# is_published = True
# birth_year = input("What is your birth year? ")
# age = 2022 - int(birth_year)
# print(age)
# print(type(age))
# print(type(birth_year))
# weight_pound = input("What is your weight in pounds? ")
# weight_kg = float(weight_pound) * 0.453592
# print("You are " + str(weight_kg) + " in kg")
# course = '''
# Hi Hesam
# Here is our first email
#
# regards
# '''
# course = 'python for beginners'
# print(course[-9])
# print(course[0:3])
# print(course[7:])
# print(course[:5])
# print(course[:])
# name = 'Jennifer'
# print(name[1:-1])
# first='John'
# last = 'Smith'
# msg = f'{first}  [{last}]  is a coder'
# print(msg)
# # print("You are " + first + last)
# print(len(first))
# print(first.upper())
# print(first.lower())
# print(first.find('J'))
# print(first.find('h'))
# print(first.find('q')) # can't find it, returns -1
# print(msg.find('is a '))
# print(msg.replace('coder','dancer'))
# print(msg.replace('Coder','dancer')) #Couldn't find Coder
# print('John' in msg)
# print('john' in msg) # Couldn't find john
# print("hesam moosapour".title())
#
# print(10 / 3)
# print(10 // 3)
# print(10 % 3)
# print(10 ** 3)
# x = 2.9
# print(round(x))
# x = -2.9
# print(abs(x))
#

# import math
# y = 2.9
# print(math.ceil(y))
# print(math.floor(y))


# is_hot = False
# is_cold = False

# print("it is a hot day"
     #       "sdfsd")

# if is_hot:
#     print("it is a hot day")
#     print("drink plenty of water")
# elif is_cold:
#     print("it is a cold day")
#     print("wear warm clothes")
# else:
#     print("it is a lovely day")
# print("enjoy your day")

# credit = 1200000
# price = 1000000
#
# has_good_credit = True
# if has_good_credit:
#     down_payment  = .1 * price
# else:
#     down_payment  = .2 * price
# print(f"Down payment is :  ${down_payment} " )
#
# if credit > price:
#     discounted_price  = price - .1 * price
# else:
#     discounted_price  = price - .2 * price
# print(f"Discounted price is :  ${discounted_price} " )

# has_high_income = True
# has_high_credit = False
# has_criminal_record = True
#
# # if has_high_income or has_high_credit:
# #     print('eligibile for loan')
# if has_high_income and not has_criminal_record:
#     print('eligibile for loan')

# weight = float(input("Weight? "))
# unit = input("(L)bs or (K)g: ")
# if unit.upper() == 'L' :
#     weight_kg = weight * 0.453592
#     print(f"You are {weight_kg} kilograms")
# elif unit.upper() == 'K' :
#     # weight_lbs = weight * 2.205
#     weight_lbs = weight / 0.453592
#     print(f"You are {weight_lbs} pounds")

# i=1
# while i<=5:
#     print('*' * i)
#     i += 1
# print("Done")
#
# secret_number = 9
# guess_count = 0
# guess_limit = 3
# while guess_count < guess_limit:
#     guess = int(input('Guess: '))
#     guess_count += 1
#     if guess == secret_number:
#         print("You won!")
#         break
# else:
#     print("Sorry, you failed!")

# if 'Help' == "HELP".capitalize():
#     print("HELP".capitalize())
# HELP.capitalize() is actually equal to Help
# exit()

#
# condition = ""
# # started = False
# while True:
#     command = input('>').capitalize()
#     if command == 'Help':
#         print('''
# start - to start the car
# stop - to stop the car
# quit - to exit
#         ''')
#     elif command == 'Start':
#         if condition == 'started':
#             # if started:
#             print('Car has already been started')
#         else:
#             condition = 'started'
#             # started = True
#             print('Car started')
#     elif command == 'Stop':
#         if condition == 'stopped':
#             #if not started:
#             print('Car has already been stopped')
#         else:
#             condition = 'stopped'
#             # started = False
#             print('Car stopped')
#     elif command == 'Quit':
#         break
#         # exit()
#     else:
#         print("I don't understand that...")


# for item in 'Python':
#     print(item)
# for item in ['Mosh','John',"Sarah"]:
#     print(item)
# for item in range(10):
#     print(item)
# for item in range(5,10):
#     print(item)
# for item in range(5, 10,2):
#     print(item)

# prices = [10,20,30]
# total_cost = 0
# for price in prices:
#     total_cost = price + total_cost
# print(total_cost)

# for x in range(4):
#     for y in range(3):
#         print(f"({x},{y})")

# numbers = [5,2,5,2,2]
# for number in numbers:
#     print('x' * number)

# x_count = 0
# for number in numbers:
#     for x_count in numbers:
#         if x_count == 5:
#             print('xxxxx')
#         elif x_count == 2:
#             print('xx')
#     break

# numbers = [5,2,5,2,7]
# for x_count in numbers:
#     output = ''
#     for count in range(x_count):
#         output += 'x'
#     print(output)

# names = ['john','bob','mosh','sarah','mary']
# print(names[:2])
# print(names[2:4])

# list = [-3,3,4,654,342,2131,44,666,-2212,4535,343]
# max = 0
# min = 0
# # max = numbers[0]
# for item in list:
#     if item > max:
#         max = item
#     if item < min:
#         min = item
# print(max)
# print(min)


#2D list
# matrix = [
#     [1,2,3],
#     [4,5,6],
#     [7,8,9]
# ]
# for row in matrix:
#     for item in row:
#         print(item)

# numbers = [5,6,34,45,2,5]
# numbers.append(22)
# print(numbers)
# numbers.insert(0,10)
# print(numbers)
# numbers.remove(34)
# print(numbers)
# # numbers.clear()
# # print(numbers)
# # numbers.pop()
# # print(numbers)
# print(numbers.index(5))
# print(50 in numbers)
# print(numbers.count(5))
# numbers.sort()
# numbers.reverse()
# print(numbers)
# numbers2 = numbers.copy()
# print(numbers2)

# numbers  = [34,4,5,2,12,54,66,4,2]
# uniques = []
# for number in numbers:
#     if number not in uniques:
#         uniques.append(number)
# print(uniques)

# tuple  = (2,4,5)
# tuple.
# You can't add anything / modify tuples

#unpacking
# coordinates = (3 , 5 ,2)
# # coordinates = [3 , 5 ,2]
# x = coordinates[0]
# y = coordinates[1]
# z = coordinates[2]
# # is the same thing as below
# x , y , z = coordinates
# print(z)

#Dictionary
# customer = {
#     'name':'john smith',
#     'age': '30',
#     'is_verified': True
# }
# print(customer['name'])
# # print(customer['Name']) # error
# print(customer.get('birthdate')) # none
# print(customer.get('birthdate','jan 20 1997')) #'jan 20 1997'

# dictionary_num = {
#     '1':'one',
#     '2': 'two',
#     '3': 'three'
# }
# phone  = input('phone: ')
# output = ''
# for item in phone:
#     if item in dictionary_num:
#         # print(dictionary_num[item])
#         output += dictionary_num.get(item) + ' '
# print(output)
#
# output = ''
# for ch in phone:
#     output += dictionary_num.get(ch,'!') + ' '
# print(output)

# message = input('>')
# words = message.split(' ')
# emojis = {
#     ':)': '😀',
#     ':(': '😞'
# }
# output = ''
# for word in words:
#     output += emojis.get(word, word) + ' '
#     # output += emojis.get(word, 'stay beastly') + ' '
# print(output)
#
# def greet_user(first_name,last_name):
#     print(f'howdy {first_name} {last_name} !')
#     print('welcome aboard')
#
# print('start')
# # greet_user('john','smith')
# # greet_user('mary')
# # greet_user(first_name='john',last_name='smith')
# greet_user('john',last_name='smith')
# # greet_user(last_name='smith', 'john') #error
# # always put keyword argument after positional argument
# print('stop')

# def square(number):
#     return number * number
#
# def square2(number):
#      print(number * number)
# print(square2(4)) #none : because it didn't return anything

# def emoji_converter(message):
#     words = message.split(' ')
#     emojis = {
#         ':)': '😀',
#         ':(': '😞'
#     }
#     output = ''
#     for word in words:
#         output += emojis.get(word, word) + ' '
#     return output
#
# print(emoji_converter(input('>')))

# try:
#     age = int(input('Age : '))
#     income = 200
#     risk = income / age
#     print(age)
# except ZeroDivisionError:
#     print('Age can\'t be 0.')
# except ValueError:
#     print('Invalid value')

# class Point:
#     def move(self):
#         print('move')
#     def drive(self):
#         print('drive')
#
# point1 = Point()
# point1.x = 10
# point1.y = 20
# print(point1.x)
# point1.drive()
#
# point2 = Point
# point2.x = 1
# print(point2.x)


# constructor
# class Point:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#
# point = Point(10,20)
# print(point.x)

# class Person:
#     def __init__(self, name):
#         self.name = name
#     def talk(self):
#         print(f'Hi, I am {self.name}')
#
# ali =  Person('alysandro')
# ali.talk()
# # print(ali.name)
# bob = Person('Bob Schneider')
# bob.talk()

# class Mammal:
#     def walk(self):
#         print('walk')
#
# class Lion(Mammal):
#     pass
#
# class Dog(Mammal):
#     def bark(self):
#         print('bark')
#
# class Cat(Mammal):
#     def be_annoying(self):
#         print('annoying')
#
# german_sheperd = Dog()
# german_sheperd.walk()
# kitty = Cat()
# kitty.be_annoying()

# Moduels
# import converters
# print(converters.lbs_to_kg(45))
# from converters import kg_to_lbs
# print(kg_to_lbs(22))

# from utils import find_max
# numbers = [12,34,77,22]
# print(find_max(numbers))
# print(max(numbers))
#
# # max = find_max(numbers)
# # print(max(numbers)) #error because max is a method
#
# maximum = find_max(numbers)
# print(maximum)

# import ecommerce.shipping
# ecommerce.shipping.calc_shipping()

# from ecommerce.shipping import calc_shipping
# calc_shipping()

# from ecommerce import shipping
# shipping.calc_shipping()

# package -> module -> class ->method

#built-in modules
# import random
# for i in range(3):
#     # print(random.random())
#     print(random.randint(10,20))

# import random
# members = ['john','mary','sarah','mosh']
# leader = random.choice(members)
# print(leader)

# import random
#
#
# class Dice:
#     def roll(self):
#         x = random.randint(1,6)
#         y = random.randint(1,6)
#         # it assumes it is a tuple
#         return x , y
#         # return (x , y)
#
# take_chance = Dice()
# print(take_chance.roll())

from pathlib import Path
# path = Path('ecommerce')
# print(path.exists())
# # path = Path('emails')
# # print(path.mkdir())
# path = Path('emails')
# print(path.rmdir())
path = Path()
# print(path.glob('*'))
# print(path.glob('*.*'))
# print(path.glob('*.py'))

# for file in path.glob('*.py'):
#     print(file)


import openpyxl as xl
from openpyxl.chart import BarChart , Reference
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
# cell = sheet['a1']
# cell1 = sheet.cell(1,1) # cell transaction_id
# print(cell1.value)
# print(sheet.max_row)
# for row in range(1,sheet.max_row + 1):
#     print(row)

for row in range(2,sheet.max_row + 1):
    cell = sheet.cell(row,3)
    # print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4) # create a column
    corrected_price_cell.value = corrected_price

values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
wb.save('transactions2.xlsx')